from selenium import webdriver
from selenium.webdriver.common.keys import Keys             # simuate keystrokes
from selenium.webdriver.common.by import By                 # import the By thing idk
from selenium.webdriver.support.ui import Select            # to select dropdown
from selenium.webdriver.support.wait import WebDriverWait   # wait for dropdown to appear
from selenium.webdriver.support import expected_conditions as EC
import time
import requests         # handle POST requests
import pandas as pd
import math
from openpyxl import load_workbook
from pathlib import Path


# ---PREPARE LIST OF LISTS AND FUNCTIONS FOR DATA---

inbox_list = []     # to store all items on search page results
headers = ["Kode", "Produk", "Nomor Registrasi", "Terbit", "Nama Produk", "Merk",
           "Kemasan", "Pendaftaran", "Alamat Pendaftaran"]

# count how many lines there are on a page
# @param input: list made by splitting page text at \n
# @param type: whether input is from 'page' or elsewhere
# @return: number of lines on a page
def num_lines(input, in_type):
    count = 0
    for line in input.split("\n"):
        count += 1
    if in_type == "page":
        print(math.floor(count / 8))
        return math.floor(count / 8)   # because each page item has 8 entries
    else:
        print(math.floor(count))
        return math.floor(count)   # because each page item has 8 entries

# move page's data into a list, then save it in an excel sheet
# @param input: list made by splitting page text at \n
def to_list(input, codes):
    inner_list = []
    count = 0
    index = 1
    for line in input.split("\n"):
        if count == 8:      # all columns are filled
            print(index)
            print(inner_list)
            inbox_list.append(inner_list)
            print()
            inner_list = []
            count = 0
            index += 1
        elif count == 2:    # strip 'Tertib: '
            line = line.strip("Tertib: ")
        elif count == 4:    # strip 'Merk: '
            line = line.strip("Merk: ")
        elif count == 5:    # strip 'Kemasan: '
            line = line.strip("Kemasan: ")

        # add item code to inner_list first
        if count == 0:    
            inner_list.append(codes[index-1])
            # product_code.append(line)

        inner_list.append(line)
        count += 1

    # append the last item in input to return_list
    print(index)
    print(inner_list)
    inbox_list.append(inner_list)
    print()

    # append to existing excel sheet
    df = pd.DataFrame(inbox_list, columns=headers)
    writer = pd.ExcelWriter("produk_vitD3.xlsx", engine="openpyxl", mode="a", if_sheet_exists="replace")
    # df.to_excel(writer, sheet_name="Produk Vitamin D3", index=True)
    df.to_excel(writer, sheet_name="Produk Vitamin D3", index=False)
    writer.close()

    print("---END---")


# --- SET UP AND NAVIGATING PAGES ---

# Selenium v4.6.0 or above: no need to set driver path
driver = webdriver.Chrome()
session = requests.Session()

url = "https://cekbpom.pom.go.id/search_home_produk"
driver.get(url)

res = session.get(url, verify=False)    # set session for POST requests later
    
# searching for search bar, then clicking it
driver.find_element(By.CLASS_NAME, "form-control").click()
time.sleep(5)

# wait for dropdown to appear, then click it + select option
WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.ID, "st_filter"))).click()
select = Select(driver.find_element(By.ID, "st_filter"))
time.sleep(5)
select.select_by_visible_text("Nama Produk")
driver.find_element(By.ID, "st_filter").click()     # click off dropdown to close it

# input text in search bar and hit enter to search
time.sleep(5)
search = driver.find_element(By.ID, "input_search")
search.send_keys("vitamin d3")
time.sleep(5)
search.send_keys(Keys.RETURN)

# get search results (that are requested with POST)
payload = {
    "st_filter": "2",
    "input_search": "vitamin d3",
    "from_home_flag": "Y"
}
res = session.post(url, data=payload, verify=False)

# store number of entries
time.sleep(5)
data_per_page = driver.find_element(By.CLASS_NAME, "kt-inbox__perpage").text.split(" ")
print(data_per_page)
next_prev = int(data_per_page[2])    # how many results per page
entries = int(data_per_page[4])      # total number of results


# check if SK and KO sheet already exists: if no, main pages haven't been parsed
xlsx = Path("produk_vitD3.xlsx")
if xlsx.is_file():
    df = pd.read_excel(xlsx, sheet_name='Produk Vitamin D3')
    print(df)
    item_codes = df["Kode"]
    print()
if not xlsx.is_file() or item_codes.size != entries:


# --- FIRST RESULT PAGE ---

    # create an excel sheet with headers
    df = pd.DataFrame(columns=headers)
    writer = pd.ExcelWriter("produk_vitD3.xlsx", engine="openpyxl")
    df.to_excel(writer, sheet_name="Produk Vitamin D3", index=False)
    writer.close()

    # get inbox 'preview' search results, make them a list of lists
    inbox = driver.find_element(By.ID, "inbox-list")

    lines = num_lines(inbox.text, "page")   # get number of items on a page
    codes = []                              # get code for each item
    for i in range(lines):
        kode = driver.find_element(By.XPATH, '//*[@id="inbox-list"]/div[{}]'.format(i+1)).get_attribute("onclick")
        kode = kode.strip("get_detail('")
        kode = kode.strip("')")
        codes.append(kode)
        # item_code.append(kode)      # this list contains all item codes

    to_list(inbox.text, codes)      # turn page contents into list


    # --- REMAINING RESULT PAGES ---

    page_no = 1
    for i in range(math.ceil(entries / next_prev) - 1):
        time.sleep(5)
        try:
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CLASS_NAME, "flaticon2-right-arrow")))
            time.sleep(5)
        finally:
            driver.find_element(By.CLASS_NAME, "flaticon2-right-arrow").click()
            WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.CLASS_NAME, "flaticon2-right-arrow")))
            print()
            print("PAGE NO: {}".format(page_no + 1))
            print()
            payload = {
                "st_filter": "2",
                "input_search": "vitamin d3",
                "from_home_flag": "Y",
                "offset": str(next_prev * page_no - 9),
                "next_prev": str(next_prev * page_no),
                "count_data_all_produk": str(entries),
                "marked": "next"
            }
            res = session.post(url, data=payload, verify=False)
            driver.implicitly_wait(20)  # set timeout limit
            inbox = driver.find_element(By.ID, "inbox-list")
            time.sleep(5)

            lines = num_lines(inbox.text, "page")   # get number of items on a page
            codes = []                              # get code for each item
            for i in range(lines):
                kode = driver.find_element(By.XPATH, '//*[@id="inbox-list"]/div[{}]'.format(i+1)).get_attribute("onclick")
                kode = kode.strip("get_detail('")
                kode = kode.strip("')")
                codes.append(kode)
                # item_code.append(kode)      # this list contains all item codes
            to_list(inbox.text, codes)      # turn page contents into list
            page_no += 1
            time.sleep(5)

    # get excel sheet ready for next part
    df = pd.read_excel(xlsx, sheet_name='Produk Vitamin D3')
    print(df)
    print()

    print("--- PAGES DONE ---")
    print()


# --- JAVASCRIPT INJECTIONS: get each item's full data ---

# get item codes and product codes from excel sheet
product_codes = df["Produk"]
item_codes = df["Kode"]

# headers for each product type
SK_headers = ["Kode", "Nomor Registrasi", "Tanggal Terbit", "Diterbitkan Oleh", "Produk", "Klaim",
              "Nama Produk", "Komposisi", "Merk", "Kemasan", "Pendaftar", "Pabrik",
              "Pengemas Primer", "Pengemas Sekunder"]

KO_headers = ["Kode", "Nomor Registrasi", "Tanggal Terbit", "Diterbitkan Oleh", "Produk",
              "Nama Produk", "Bentuk Sediaan", "Merk", "Kemasan", "Pendaftar", "1. Industri Kosmetika",]

SK_list = []
KO_list = []
SK_last = 0
KO_last = 0

# continue scraping where program left off
wb = load_workbook("produk_vitD3.xlsx", read_only=True)

# grab the last inputted code for both sheets
if "Suplemen" in wb.sheetnames:
    print("Ada sheet Suplemen")
    SK_df = pd.read_excel(xlsx, 'Suplemen')     # sheet of SK items and details
    SK_curr_codes = SK_df["Kode"]
    SK_last = str(SK_curr_codes.iloc[-1])
    print("SK_last: {}".format(SK_last))

if "Kosmetika" in wb.sheetnames:
    print("Ada sheet Kosmetika")
    KO_df = pd.read_excel(xlsx, 'Kosmetika')    # sheet of KO items adn details
    KO_curr_codes = KO_df["Kode"]
    KO_last = str(KO_curr_codes.iloc[-1])
    print("KO_last: {}".format(KO_last))

# go through item list to find last SK and KO item scraped
code_index = 0
SK_index = 0
KO_index = 0
start_index = 0

for line in item_codes:
    # code = str(item_codes[code_index])
    if line == SK_last:
        SK_index = code_index
        print("SK found")
    if line == KO_last:
        KO_index = code_index
        print("KO found")
    code_index += 1

# determine which item is more recent; set its index to starting index
if SK_index > KO_index:
    start_index = SK_index
else:
    start_index = KO_index

if start_index != 0:    # to avoid the last written entry to be written again
    start_index += 1

item_count = start_index

print(start_index)
print()

for i in range(start_index, entries):    # iterate through each item via injection
    result = ""
    inner_list = []
    # item_count += 1

    code = item_codes[i]    # BUG: ValueError: 0 is not in range? only when start from empty sheet
    injection = 'get_detail("' + code + '")'
    product = product_codes[i]

    driver.execute_script(injection)    # to open detail window
    time.sleep(10)
    
    
    # scrape details, save into list
    WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="detailobat"]/table/tbody')))
    obat = driver.find_element(By.XPATH, '//*[@id="detailobat"]/table/tbody')
    data = obat.text.split("\n")
    print(i)
    print(injection)
    print(product)
    data_lines = num_lines(obat.text, " ")   # get number of lines
    print()

    driver.execute_script(injection)    # to close detail window
    time.sleep(8)

    # append to existing sheet
    writer = pd.ExcelWriter("produk_vitD3.xlsx", engine="openpyxl", mode="a", if_sheet_exists="overlay")

    inner_list.append(code)     # add item code to sheet

    # sort data into list of lists first
    if product == "SK":
        col = 1             # start at second column, since first is code
        last_col = False
        col_name = ""
        line_count = 0
        for line in data:
            line_count += 1

            if line == "Pendaftar & Importir":  # deal with header inconsistency. sigh.
                line = line.replace("Pendaftar & Importir", "Pendaftar")
                print(line)
            
            if line == "Pengemas Primer":       # deal with even MORE inconsistencies. sigh.
                last_col = False
            if line == "Pengemas Sekunder":
                last_col = True

            if line == data[0]:     # ignore first line of data, which is a header
                col += 1
                # col_name = line
                print(line)
                continue
            elif (line != data[0]) and (col < len(SK_headers)) and (line == SK_headers[col]):
                print("   ^ Inserting into column: {}".format(result))
                print(line)
                inner_list.append(result)       # sort data into the previously read column
                result = ""
                col += 1
                col_name = line
                if line == "Pabrik":    # hardcode check the last column
                    last_col = True
                continue
            else:   # if not a header, store each column entry as a string
                result += line

            if data_lines == line_count and last_col:   # deal with last entry
                print("  Inserting last entry: {}".format(result))
                inner_list.append(result)
        
        # deal with the extra columns only a few items have. i hate this inconsistency. i hate it
        if len(inner_list) == 12:
            inner_list.append("-")
            inner_list.append("-")

        # add data to list of lists, write to excel
        SK_list.append(inner_list)
        
    elif product == "KO":
        col = 1             # start at second column, since first is code
        last_col = False
        col_name = ""
        line_count = 0
        for line in data:
            line_count += 1

            if line == data[0]:     # ignore first line of data, which is a header
                col += 1
                # col_name = line
                print(line)
                continue
            elif (line != data[0]) and (col < len(KO_headers)) and (line == KO_headers[col]):
                print("   ^ Inserting into column: {}".format(result))
                print(line)
                inner_list.append(result)       # sort data into the previously read column
                result = ""
                col += 1
                col_name = line
                if line == "1. Industri Kosmetika":    # hardcode check the last column
                    last_col = True
                continue
            else:   # if not a header, store each column entry as a string
                result += line

            if data_lines == line_count and last_col:   # deal with last entry
                print("  Inserting last entry: {}".format(result))
                inner_list.append(result)
        
        # add data to list of lists, write to excel
        KO_list.append(inner_list)

    print()
    for j in range(len(inner_list)):
        print ("{}: {}".format(j + 1, inner_list[j]))


    # write to excel every 5 entries, or when all items are read
    if item_count != 0 and (item_count % 5 == 0 or item_count == entries):
        df_sk = pd.DataFrame(SK_list, columns=SK_headers)
        df_ko = pd.DataFrame(KO_list, columns=KO_headers)

        if start_index == 0:    # if sheets are just created, keep headers + max_row won't work
            df_sk.to_excel(writer, sheet_name="Suplemen", index=False)
            df_ko.to_excel(writer, sheet_name="Kosmetika", index=False)
        else:
            df_sk.to_excel(writer, sheet_name="Suplemen", startrow=SK_index, header=None, index=False)
            df_ko.to_excel(writer, sheet_name="Kosmetika", startrow=KO_index, header=None, index=False)
        
        print("--WROTE TO SHEET")
        time.sleep(10)

    writer.close()
    print("DONE: item {}".format(i+1))
    item_count += 1
    print()

print("PROGRAM DONE")

driver.quit()
